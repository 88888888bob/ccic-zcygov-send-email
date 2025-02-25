from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service # 导入 Service 类
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl

def clickXPATH(xpath):
    button = driver.find_element(By.XPATH,xpath)
    button.click()

def inputTextXPATH(xpath,text):
    text_label  = driver.find_element(By.XPATH,xpath)
    text_label.send_keys(text)

def clearTextXPATH(xpath):
    text_label  = driver.find_element(By.XPATH,xpath)
    text_label.clear()


send=False

mails=[]
with open("mail.txt","r+") as file:
   mail=file.read().split("\n")
for m in mail:
    if m!="":
        mails.append(m)

data = openpyxl.Workbook()

sheet = data["Sheet"]

#filepath="temp "+generate_code(5)+" .xlsx"
filepath3="浙江政府采购网 "+nowTime+"生成 .xlsx"
data.save(filepath3)
filepath=filepath3

# Chrome options for running headless (no GUI)
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# ChromeDriver 的路径 (使用 apt 安装时，通常在这里)
chrome_driver_path = "/usr/lib/chromium-browser/chromedriver"

# 创建 Service 对象，指定 ChromeDriver 的路径
service = Service(executable_path=chrome_driver_path)

# 初始化 Chrome WebDriver，并传入 Service 对象
driver = webdriver.Chrome(service=service, options=chrome_options)

try:
    driver.get('https://zfcg.czt.zj.gov.cn/site/home')
    time.sleep(0.1)


    clickXPATH("/html/body/div/div/div[3]/div/div/ul/a[3]/li")
    time.sleep(1)
    clickXPATH("/html/body/div/div/div[4]/div/div[2]/div[1]/div/div/div[4]/div[1]/div/span[2]")
    time.sleep(1)
    clickXPATH("/html/body/div/div/div[4]/div/div[2]/div[2]/div[1]/div/div/form/div[9]/div/button[3]/span/span")
    #time.sleep(2)
    #clickXPATH("/html/body/div/div/div[4]/div/div[2]/div[2]/div[1]/div/div/form/div[9]/div/button[3]/span/span")
    time.sleep(1)
    inputTextXPATH("/html/body/div/div/div[4]/div/div[2]/div[2]/div[1]/div/div/form/div[1]/div/div/input","保险")
    time.sleep(1.5)
    clickXPATH("/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[1]/div/div/form/div[8]/div/div[2]/div/input")
    time.sleep(1.5)
    clickXPATH("/html/body/div[2]/div/div[1]/ul/li[5]")
    time.sleep(1.5)
    clickXPATH("/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[1]/div/div/form/div[9]/div/button[1]")
    time.sleep(3)
    page=driver.find_elements(By.XPATH, '/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[2]/div/div[1]/ul/li')

    sheet=data.active
        
    sheet.cell(1,1).value="项目地点"
    sheet.cell(1,2).value="项目名称"
    sheet.cell(1,3).value="发布日期"
    sheet.cell(1,4).value="项目链接"
    maxLenTitle=0
    maxLenUrl=0
    lines=2
    #初始化变量，无需更改

    getPages=1
    #要爬取的页数

    for i in range(getPages):
        page=driver.find_elements(By.XPATH, '/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[2]/div/div[1]/ul/li')
        for li in page:
            link=li.find_element(By.XPATH, 'a')
            day=li.find_element(By.XPATH, 'span[2]')
            
            position=li.find_element(By.XPATH, 'span[1]')

            title=link.text
            url=link.get_attribute('href')
            day=day.text
            position=position.text.replace("\n","")

            
            if maxLenTitle<len(title):
                maxLenTitle=len(title)
            if maxLenUrl<len(url):
                maxLenUrl=len(url)

            sheet.cell(lines,1).value=position
            sheet.cell(lines,2).value=title
            sheet.cell(lines,3).value=day
            sheet.cell(lines,4).hyperlink=url
            if nowTime==day:
                sendDatas.append([position,title,day,url])
            
            lines+=1
        clickXPATH("/html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[2]/div/div[2]/div/button[2]")
        #           /html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[2]/div/div[2]/div/button[2]   
        #           /html/body/div[1]/div/div[4]/div/div[2]/div[2]/div[2]/div/div[2]/div/button[2]
    sheet.column_dimensions["A"].width=25
    sheet.column_dimensions["B"].width=maxLenTitle*2
    sheet.column_dimensions["C"].width=15
    sheet.column_dimensions["D"].width=maxLenUrl

    data.save(filepath)

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    driver.quit()
    print("Browser closed.")
